from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import load_workbook

import journal_metrics


def _args(**overrides: object) -> SimpleNamespace:
    values = {
        "adapter": "sealib",
        "input": None,
        "update": False,
        "db_path": "/tmp/sealib.test.sqlite",
        "country": None,
        "sinta_command": None,
        "sinta_python": None,
        "sinta_timeout": 180,
    }
    values.update(overrides)
    return SimpleNamespace(**values)


def _sealib_envelope(
    *,
    external_journal_id: str = "(01)001IO00001",
    issn: str | None = "1234-5678",
    note: str | None = None,
) -> dict:
    return {
        "status": "fetched",
        "source": "SEALIB",
        "query": "Journal Name",
        "candidates": [
            {
                "source": "SEALIB",
                "external_journal_id": external_journal_id,
                "title": "Journal Name",
                "issn": issn,
                "eissn": None,
                "publisher": None,
                "country": "IO",
                "grade": None,
                "url": None,
                "note": note,
            }
        ],
        "error": None,
    }


def _sinta_envelope(
    *,
    status: str = "fetched",
    title: str = "Journal Name",
    grade: str | None = "S1 Accredited",
) -> dict:
    candidates = []
    if status in {"fetched", "multiple_candidates"}:
        candidates = [
            {
                "source": "SINTA",
                "external_journal_id": "SINTA-NEW",
                "title": title,
                "issn": None,
                "eissn": None,
                "publisher": "Publisher",
                "country": "ID",
                "grade": grade,
                "url": "https://example.test/sinta-new",
                "note": None,
            }
        ]
    return {
        "status": status,
        "source": "SINTA",
        "query": title,
        "candidates": candidates,
        "error": "adapter failed" if status == "adapter_error" else None,
    }


class FetchJournalCommandTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.workbook_path = Path(self.temp_dir.name) / "journal_metrics.xlsx"

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def _save_template_with_main_row(self, values: dict[str, object]) -> None:
        self._save_template_with_main_rows([values])

    def _save_template_with_main_rows(self, rows: list[dict[str, object]]) -> None:
        wb = journal_metrics.build_template_workbook()
        main = wb["main"]
        for values in rows:
            row = [values.get(header) for header in journal_metrics.MAIN_HEADERS]
            main.append(row)
        wb.save(self.workbook_path)

    def _append_journal_row(self, values: dict[str, object]) -> None:
        wb = load_workbook(self.workbook_path)
        journal = wb["journal"]
        journal.append([
            values.get(header)
            for header in journal_metrics.JOURNAL_HEADERS
        ])
        wb.save(self.workbook_path)

    def _append_convert_row(self, values: dict[str, object]) -> None:
        wb = load_workbook(self.workbook_path)
        convert = wb["convert"]
        convert.append([
            values.get(header)
            for header in journal_metrics.CONVERT_HEADERS
        ])
        wb.save(self.workbook_path)

    def test_sealib_fetch_fills_main_id_issn_and_keeps_status_pending(self) -> None:
        self._save_template_with_main_row(
            {
                "name": "Journal Name",
                "status": "pending",
            }
        )

        with patch(
            "journal_metrics.fetch_sealib_journal",
            return_value=_sealib_envelope(
                external_journal_id="SEALIB-001",
                issn="0024-9521",
                note=None,
            ),
        ):
            journal_metrics.fetch_journal_command(
                _args(input=str(self.workbook_path))
            )

        wb = load_workbook(self.workbook_path)
        main = wb["main"]
        headers = journal_metrics.header_index(main)
        self.assertEqual(main.cell(row=2, column=headers["id"] + 1).value, "SEALIB-001")
        self.assertEqual(main.cell(row=2, column=headers["issn"] + 1).value, "0024-9521")
        self.assertIsNone(main.cell(row=2, column=headers["o_name"] + 1).value)
        self.assertEqual(main.cell(row=2, column=headers["status"] + 1).value, "pending")

    def test_sealib_fetch_does_not_fill_empty_candidate_issn(self) -> None:
        self._save_template_with_main_row(
            {
                "name": "Journal Name",
                "status": "pending",
            }
        )

        with patch(
            "journal_metrics.fetch_sealib_journal",
            return_value=_sealib_envelope(issn=None, note="Original Name"),
        ):
            journal_metrics.fetch_journal_command(
                _args(input=str(self.workbook_path))
            )

        wb = load_workbook(self.workbook_path)
        main = wb["main"]
        headers = journal_metrics.header_index(main)
        self.assertIsNone(main.cell(row=2, column=headers["issn"] + 1).value)
        self.assertEqual(main.cell(row=2, column=headers["o_name"] + 1).value, "Original Name")
        self.assertEqual(main.cell(row=2, column=headers["status"] + 1).value, "pending")

    def test_sinta_fetch_passes_main_issn_and_eissn_to_adapter(self) -> None:
        self._save_template_with_main_row(
            {
                "id": "SEALIB-001",
                "name": "Journal Name",
                "issn": "0024-9521",
                "eissn": "2354-9114",
                "search_query": "Indonesian Journal of Geography",
                "status": "pending",
            }
        )
        envelope = {
            "status": "multiple_candidates",
            "source": "SINTA",
            "query": "Indonesian Journal of Geography",
            "candidates": [],
            "error": None,
        }

        with patch(
            "journal_metrics.fetch_sinta_journal",
            return_value=envelope,
        ) as fetch_sinta:
            journal_metrics.fetch_journal_command(
                _args(
                    adapter="sinta",
                    input=str(self.workbook_path),
                    sinta_command="/tmp/sinta-full-cli-v3.py",
                )
            )

        _positional_args, keyword_args = fetch_sinta.call_args
        self.assertEqual(keyword_args["main_issn"], "0024-9521")
        self.assertEqual(keyword_args["main_eissn"], "2354-9114")

    def test_without_update_does_not_reprocess_terminal_statuses(self) -> None:
        self._save_template_with_main_rows(
            [
                {"search_query": "Fetched", "status": "fetched"},
                {"search_query": "Not Found", "status": "not_found"},
                {"search_query": "Multiple", "status": "multiple_candidates"},
                {"search_query": "Pending", "status": "pending"},
            ]
        )

        with patch(
            "journal_metrics.fetch_sinta_journal",
            return_value=_sinta_envelope(title="Pending"),
        ) as fetch_sinta:
            journal_metrics.fetch_journal_command(
                _args(
                    adapter="sinta",
                    input=str(self.workbook_path),
                    sinta_command="/tmp/sinta-full-cli-v3.py",
                )
            )

        self.assertEqual(fetch_sinta.call_count, 1)
        self.assertEqual(fetch_sinta.call_args.args[0], "Pending")

    def test_update_reprocesses_terminal_statuses_but_not_skip_or_done(self) -> None:
        self._save_template_with_main_rows(
            [
                {"search_query": "Fetched", "status": "fetched"},
                {"search_query": "Not Found", "status": "not_found"},
                {"search_query": "Multiple", "status": "multiple_candidates"},
                {"search_query": "Skip", "status": "skip"},
                {"search_query": "Done", "status": "done"},
            ]
        )

        with patch(
            "journal_metrics.fetch_sinta_journal",
            return_value=_sinta_envelope(),
        ) as fetch_sinta:
            journal_metrics.fetch_journal_command(
                _args(
                    adapter="sinta",
                    input=str(self.workbook_path),
                    update=True,
                    sinta_command="/tmp/sinta-full-cli-v3.py",
                )
            )

        self.assertEqual(fetch_sinta.call_count, 3)
        self.assertEqual(
            [call.args[0] for call in fetch_sinta.call_args_list],
            ["Fetched", "Not Found", "Multiple"],
        )

    def test_should_process_update_processes_all_statuses_except_skip_or_done(self) -> None:
        process_statuses = [
            "",
            "pending",
            "adapter_error",
            "fetched",
            "not_found",
            "multiple_candidates",
            "hold",
        ]
        for status in process_statuses:
            with self.subTest(status=status):
                self.assertTrue(
                    journal_metrics.should_process_main_row(
                        status,
                        "sealib",
                        update=True,
                    )
                )

        for status in ["skip", "done"]:
            with self.subTest(status=status):
                self.assertFalse(
                    journal_metrics.should_process_main_row(
                        status,
                        "sealib",
                        update=True,
                    )
                )

    def test_sinta_update_replaces_existing_sinta_rows_for_same_main_row_only(self) -> None:
        self._save_template_with_main_row(
            {
                "search_query": "Journal Name",
                "status": "fetched",
            }
        )
        self._append_journal_row(
            {
                "main_row_id": 2,
                "journal_type": "SINTA",
                "external_journal_id": "SINTA-OLD",
                "journal_name": "Journal Name",
                "grade": "S2 Accredited",
                "fetch_status": "ok",
            }
        )
        self._append_journal_row(
            {
                "main_row_id": 2,
                "journal_type": "SEALIB",
                "external_journal_id": "SEALIB-KEEP",
                "journal_name": "Journal Name",
                "fetch_status": "ok",
            }
        )
        self._append_journal_row(
            {
                "main_row_id": 2,
                "journal_type": "MOCK",
                "external_journal_id": "MOCK-KEEP",
                "journal_name": "Journal Name",
                "fetch_status": "ok",
            }
        )
        self._append_journal_row(
            {
                "main_row_id": 99,
                "journal_type": "SINTA",
                "external_journal_id": "SINTA-OTHER",
                "journal_name": "Other",
                "grade": "S3 Accredited",
                "fetch_status": "ok",
            }
        )

        with patch(
            "journal_metrics.fetch_sinta_journal",
            return_value=_sinta_envelope(grade="S1 Accredited"),
        ):
            journal_metrics.fetch_journal_command(
                _args(
                    adapter="sinta",
                    input=str(self.workbook_path),
                    update=True,
                    sinta_command="/tmp/sinta-full-cli-v3.py",
                )
            )

        wb = load_workbook(self.workbook_path)
        journal = wb["journal"]
        headers = journal_metrics.header_index(journal)
        rows = list(journal.iter_rows(min_row=2, values_only=True))
        row_dicts = [
            journal_metrics.row_to_dict(list(headers.keys()), row)
            for row in rows
        ]
        self.assertNotIn(
            "SINTA-OLD",
            [row.get("external_journal_id") for row in row_dicts],
        )
        self.assertIn(
            "SEALIB-KEEP",
            [row.get("external_journal_id") for row in row_dicts],
        )
        self.assertIn(
            "MOCK-KEEP",
            [row.get("external_journal_id") for row in row_dicts],
        )
        self.assertIn(
            "SINTA-OTHER",
            [row.get("external_journal_id") for row in row_dicts],
        )
        new_sinta_rows = [
            row
            for row in row_dicts
            if row.get("main_row_id") == 2 and row.get("journal_type") == "SINTA"
        ]
        self.assertEqual(len(new_sinta_rows), 1)
        self.assertEqual(new_sinta_rows[0]["external_journal_id"], "SINTA-NEW")
        self.assertEqual(new_sinta_rows[0]["grade"], "S1 Accredited")

    def test_sinta_update_adapter_error_deletes_old_sinta_rows_and_adds_no_error_row(self) -> None:
        self._save_template_with_main_row(
            {
                "search_query": "Journal Name",
                "status": "fetched",
            }
        )
        self._append_journal_row(
            {
                "main_row_id": 2,
                "journal_type": "SINTA",
                "external_journal_id": "SINTA-OLD",
                "journal_name": "Journal Name",
                "fetch_status": "ok",
            }
        )

        with patch(
            "journal_metrics.fetch_sinta_journal",
            return_value=_sinta_envelope(status="adapter_error"),
        ):
            journal_metrics.fetch_journal_command(
                _args(
                    adapter="sinta",
                    input=str(self.workbook_path),
                    update=True,
                    sinta_command="/tmp/sinta-full-cli-v3.py",
                )
            )

        wb = load_workbook(self.workbook_path)
        main = wb["main"]
        main_headers = journal_metrics.header_index(main)
        self.assertEqual(main.cell(row=2, column=main_headers["status"] + 1).value, "adapter_error")
        journal = wb["journal"]
        self.assertEqual(journal.max_row, 1)

    def test_sealib_update_replaces_existing_sealib_rows_for_same_main_row_only(self) -> None:
        self._save_template_with_main_row(
            {
                "name": "Journal Name",
                "status": "fetched",
            }
        )
        self._append_journal_row(
            {
                "main_row_id": 2,
                "journal_type": "SEALIB",
                "external_journal_id": "SEALIB-OLD",
                "journal_name": "Journal Name",
                "fetch_status": "ok",
            }
        )
        self._append_journal_row(
            {
                "main_row_id": 2,
                "journal_type": "SINTA",
                "external_journal_id": "SINTA-KEEP",
                "journal_name": "Journal Name",
                "grade": "S1 Accredited",
                "fetch_status": "ok",
            }
        )
        self._append_journal_row(
            {
                "main_row_id": 2,
                "journal_type": "MOCK",
                "external_journal_id": "MOCK-KEEP",
                "journal_name": "Journal Name",
                "fetch_status": "ok",
            }
        )
        self._append_journal_row(
            {
                "main_row_id": 99,
                "journal_type": "SEALIB",
                "external_journal_id": "SEALIB-OTHER",
                "journal_name": "Other",
                "fetch_status": "ok",
            }
        )

        with patch(
            "journal_metrics.fetch_sealib_journal",
            return_value=_sealib_envelope(external_journal_id="SEALIB-NEW"),
        ):
            journal_metrics.fetch_journal_command(
                _args(input=str(self.workbook_path), update=True)
            )

        wb = load_workbook(self.workbook_path)
        journal = wb["journal"]
        headers = journal_metrics.header_index(journal)
        row_dicts = [
            journal_metrics.row_to_dict(list(headers.keys()), row)
            for row in journal.iter_rows(min_row=2, values_only=True)
        ]
        external_ids = [row.get("external_journal_id") for row in row_dicts]
        self.assertNotIn("SEALIB-OLD", external_ids)
        self.assertIn("SINTA-KEEP", external_ids)
        self.assertIn("MOCK-KEEP", external_ids)
        self.assertIn("SEALIB-OTHER", external_ids)

        new_sealib_rows = [
            row
            for row in row_dicts
            if row.get("main_row_id") == 2 and row.get("journal_type") == "SEALIB"
        ]
        self.assertEqual(len(new_sealib_rows), 1)
        self.assertEqual(new_sealib_rows[0]["external_journal_id"], "SEALIB-NEW")

    def test_sealib_update_does_not_change_status_or_overwrite_existing_main_values(self) -> None:
        self._save_template_with_main_row(
            {
                "id": "SEALIB-OLD",
                "name": "Journal Name",
                "o_name": "Original Name",
                "issn": "1111-1111",
                "status": "fetched",
            }
        )

        with patch(
            "journal_metrics.fetch_sealib_journal",
            return_value=_sealib_envelope(
                external_journal_id="SEALIB-NEW",
                issn="2222-2222",
                note="New Original Name",
            ),
        ):
            journal_metrics.fetch_journal_command(
                _args(input=str(self.workbook_path), update=True)
            )

        wb = load_workbook(self.workbook_path)
        main = wb["main"]
        headers = journal_metrics.header_index(main)
        self.assertEqual(main.cell(row=2, column=headers["id"] + 1).value, "SEALIB-OLD")
        self.assertEqual(main.cell(row=2, column=headers["issn"] + 1).value, "1111-1111")
        self.assertEqual(main.cell(row=2, column=headers["o_name"] + 1).value, "Original Name")
        self.assertEqual(main.cell(row=2, column=headers["status"] + 1).value, "fetched")

    def test_sinta_update_passes_issn_values_for_detail_selection_path(self) -> None:
        self._save_template_with_main_row(
            {
                "search_query": "Indonesian Journal of Geography",
                "issn": "0024-9521",
                "eissn": "2354-9114",
                "status": "multiple_candidates",
            }
        )

        with patch(
            "journal_metrics.fetch_sinta_journal",
            return_value=_sinta_envelope(),
        ) as fetch_sinta:
            journal_metrics.fetch_journal_command(
                _args(
                    adapter="sinta",
                    input=str(self.workbook_path),
                    update=True,
                    sinta_command="/tmp/sinta-full-cli-v3.py",
                )
            )

        _positional_args, keyword_args = fetch_sinta.call_args
        self.assertEqual(keyword_args["main_issn"], "0024-9521")
        self.assertEqual(keyword_args["main_eissn"], "2354-9114")

    def test_convert_resets_existing_convert_rows_before_regeneration(self) -> None:
        self._save_template_with_main_row(
            {
                "id": "SEALIB-001",
                "name": "Journal Name",
                "o_name": "Original Name",
                "status": "fetched",
            }
        )
        self._append_journal_row(
            {
                "main_row_id": 2,
                "journal_type": "SINTA",
                "external_journal_id": "SINTA-NEW",
                "journal_name": "Journal Name",
                "grade": "S1 Accredited",
                "profile_url": "https://example.test/sinta-new",
                "fetch_status": "ok",
                "raw_json": '{"country": "ID", "external_journal_id": "SINTA-NEW"}',
            }
        )
        self._append_convert_row(
            {
                "main_row_id": 2,
                "metric_source": "SINTA",
                "sealib_name": "Old Convert Row",
                "grade": "S2 Accredited",
                "convert_status": "ready",
            }
        )

        journal_metrics.convert_command(
            SimpleNamespace(input=str(self.workbook_path))
        )

        wb = load_workbook(self.workbook_path)
        convert = wb["convert"]
        headers = journal_metrics.header_index(convert)
        rows = [
            journal_metrics.row_to_dict(list(headers.keys()), row)
            for row in convert.iter_rows(min_row=2, values_only=True)
        ]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["sealib_name"], "Journal Name")
        self.assertEqual(rows[0]["metric_source"], "SINTA")
        self.assertEqual(rows[0]["grade"], "S1 Accredited")
        self.assertNotEqual(rows[0]["sealib_name"], "Old Convert Row")


if __name__ == "__main__":
    unittest.main()
