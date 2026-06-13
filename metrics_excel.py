#!/usr/bin/env python3
"""Program 1: build and update a Journal Metrics Excel correspondence table."""

from __future__ import annotations

import argparse
import csv
import json
import re
import shlex
import sqlite3
import subprocess
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote

from openpyxl import Workbook, load_workbook


RAW_HEADERS = [
    "query",
    "journal_name",
    "sinta_level",
    "affiliation",
    "journal_id",
    "profile_url",
    "fetched_at",
]

DETAIL_HEADERS = [
    "p_issn",
    "e_issn",
    "subject_area",
    "website_url",
    "editor_url",
    "garuda_url",
    "google_scholar_url",
]

MAP_HEADERS = [
    "confirmed",
    "status",
    "metric_source",
    "metric_country",
    "sealib_id",
    "sealib_name",
    "sealib_o_name",
    "issn",
    "sinta_query_name",
    "sinta_journal_name",
    "sinta_journal_id",
    "profile_url",
    "sinta_level",
    "garuda_name",
    "fetched_at",
]

EXPORT_HEADERS = [
    "metric_source",
    "metric_country",
    "sealib_name",
    "sealib_o_name",
    "sealib_id",
    "grade",
    "url",
    "note",
]

COUNTRY_ALIASES = {
    "ID": "IO",  # SINTA metric country is ISO-like ID; SEALIB header.country uses LC code IO.
}

DEFAULT_IMPORT_SRC_SHEET = "インドネシア (名前あり)"
DEFAULT_IMPORT_ANALYSIS_SHEET = "インドネシア（SINTAツール分析）"

LEGACY_MAP_COLUMNS = {
    "sealib_id": ["sealib_id", "id"],
    "sealib_name": ["sealib_name", "name"],
    "sealib_o_name": ["sealib_o_name", "o_name"],
    "issn": ["issn"],
    "sinta_query_name": ["sinta_query_name", "SINTAでの名前（検索可能）", "SINTAでの名前"],
    "sinta_journal_name": ["sinta_journal_name", "SINTAでの名前"],
    "sinta_journal_id": ["sinta_journal_id", "journal_id"],
    "profile_url": ["profile_url", "sinta_url", "url"],
    "sinta_level": ["sinta_level"],
    "garuda_name": ["garuda_name", "Garudaでの名前（SINTA未掲載）"],
}


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def resolve_adapter_command(args: argparse.Namespace) -> list[str]:
    """Resolve the adapter invocation tokens (script path) for fetch/refresh.

    ``--adapter-command`` (script path) is primary; ``--sinta-cli`` is a
    backward-compatible alias. The interpreter is taken from ``--python``;
    an interpreter-prefixed command (``python3 ...``) must not be passed.
    There is no default: fetch/refresh stop with a clear error when neither
    is given.
    """
    adapter = getattr(args, "adapter_command", None)
    sinta_cli = getattr(args, "sinta_cli", None)
    if adapter:
        if sinta_cli:
            print(
                "WARNING: both --adapter-command and --sinta-cli given; using --adapter-command.",
                file=sys.stderr,
            )
        tokens = shlex.split(adapter)
    elif sinta_cli:
        tokens = [str(sinta_cli)]
    else:
        raise SystemExit(
            "ERROR: --adapter-command is required for fetch/refresh "
            "(script path to the adapter, e.g. ../sinta-full-cli-v3/sinta-full-cli-v3.py). "
            "--sinta-cli remains as a backward-compatible alias. "
            "The interpreter is taken from --python."
        )
    if not tokens:
        raise SystemExit("ERROR: --adapter-command is empty.")
    script_path = Path(tokens[-1])
    if not script_path.exists():
        raise SystemExit(
            f"ERROR: adapter script not found: {script_path} "
            "(pass --adapter-command <script path>; the interpreter is taken from --python)."
        )
    return tokens


def sheet_names(source: str, country: str) -> tuple[str, str]:
    return f"{country}（{source}ツール分析）", f"{country} (名前あり)"


def load_or_create_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def ensure_sheet(wb: Workbook, title: str, headers: list[str]):
    ws = wb[title] if title in wb.sheetnames else wb.create_sheet(title)
    if ws.max_row == 1 and all(ws.cell(1, col).value is None for col in range(1, ws.max_column + 1)):
        for col, header in enumerate(headers, start=1):
            ws.cell(1, col).value = header
        ws.freeze_panes = "A2"
        return ws

    existing = [str(ws.cell(1, col).value or "").strip() for col in range(1, ws.max_column + 1)]
    if not any(existing):
        for col, header in enumerate(headers, start=1):
            ws.cell(1, col).value = header
        ws.freeze_panes = "A2"
        return ws

    for header in headers:
        if header not in existing:
            ws.cell(1, ws.max_column + 1).value = header
            existing.append(header)
    ws.freeze_panes = "A2"
    return ws


def header_index(ws) -> dict[str, int]:
    return {
        str(ws.cell(1, col).value).strip(): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(1, col).value is not None
    }


def get_value(ws, row: int, header: str) -> str:
    idx = header_index(ws).get(header)
    if idx is None:
        return ""
    value = ws.cell(row, idx).value
    return "" if value is None else str(value).strip()


def set_value(ws, row: int, header: str, value: Any) -> None:
    idx = header_index(ws).get(header)
    if idx is None:
        idx = ws.max_column + 1
        ws.cell(1, idx).value = header
    ws.cell(row, idx).value = value


def is_confirmed(value: str) -> bool:
    return value.strip() == "1"


def normalized_grade(raw: Any) -> tuple[str, str | None]:
    text = "" if raw is None else str(raw).strip()
    if text == "":
        return "", None

    lowered = text.lower()
    for level in range(1, 7):
        canonical = f"S{level} Accredited"
        if lowered == canonical.lower():
            return canonical, None
        if lowered in {f"sinta {level}", f"s{level}"}:
            return canonical, None

    return text, f"WARNING: unknown sinta_level kept as raw value: {text}"


def header_positions(ws) -> dict[str, list[int]]:
    positions: dict[str, list[int]] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value is None:
            continue
        positions.setdefault(str(value).strip(), []).append(col)
    return positions


def get_legacy_value(ws, row: int, positions: dict[str, list[int]], aliases: list[str], prefer: str = "first") -> str:
    for alias in aliases:
        cols = positions.get(alias)
        if not cols:
            continue
        col = cols[-1] if prefer == "last" else cols[0]
        value = ws.cell(row, col).value
        text = "" if value is None else str(value).strip()
        if text:
            return text
    return ""


def extract_journal_id(profile_url: str) -> str:
    match = re.search(r"/journals/profile/(\d+)|[?&]id=(\d+)", profile_url or "")
    if not match:
        return ""
    return match.group(1) or match.group(2) or ""


def row_has_content(ws, row: int) -> bool:
    return any(ws.cell(row, col).value not in (None, "") for col in range(1, ws.max_column + 1))


def read_sealib_rows(db_path: Path, country: str) -> list[sqlite3.Row]:
    resolved = db_path.resolve()
    uri = f"file:{quote(str(resolved), safe='/:')}?mode=ro"
    conn = sqlite3.connect(uri, uri=True)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """
            SELECT id, name, o_name, issn
              FROM header
             WHERE UPPER(country) = UPPER(?)
             ORDER BY id
            """,
            (country,),
        ).fetchall()
    finally:
        conn.close()
    return rows


def init_command(args: argparse.Namespace) -> int:
    wb = load_or_create_workbook(args.xlsx)
    raw_title, map_title = sheet_names(args.source, args.country)
    ensure_sheet(wb, raw_title, RAW_HEADERS)
    ws = ensure_sheet(wb, map_title, MAP_HEADERS)

    db_path = args.db_path
    sealib_country = args.sealib_country or COUNTRY_ALIASES.get(args.country.upper(), args.country.upper())
    db_rows = read_sealib_rows(db_path, sealib_country)

    existing_ids = {
        get_value(ws, row, "sealib_id")
        for row in range(2, ws.max_row + 1)
        if get_value(ws, row, "sealib_id")
    }

    added = 0
    for item in db_rows:
        if item["id"] in existing_ids:
            continue
        ws.append([
            "",
            "needs_fetch",
            args.source,
            args.country,
            item["id"] or "",
            item["name"] or "",
            item["o_name"] or "",
            item["issn"] or "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ])
        added += 1

    wb.save(args.xlsx)
    print(f"initialized: {args.xlsx}")
    print(f"sheet: {map_title}")
    print(f"source DB: {db_path} (country={sealib_country}, read-only SELECT)")
    print(f"rows found: {len(db_rows)}, added: {added}")
    return 0


def find_existing_map_rows(ws) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row in range(2, ws.max_row + 1):
        sealib_id = get_value(ws, row, "sealib_id")
        if sealib_id and sealib_id not in rows:
            rows[sealib_id] = row
    return rows


def import_existing_master_command(args: argparse.Namespace) -> int:
    src_wb = load_workbook(args.input, data_only=True)
    if args.src_sheet not in src_wb.sheetnames:
        raise ValueError(f"source sheet not found: {args.src_sheet}")
    if args.analysis_sheet not in src_wb.sheetnames:
        raise ValueError(f"analysis sheet not found: {args.analysis_sheet}")

    out_wb = load_or_create_workbook(args.xlsx)
    raw_title, map_title = sheet_names(args.source, args.country)
    raw_ws = ensure_sheet(out_wb, raw_title, RAW_HEADERS)
    map_ws = ensure_sheet(out_wb, map_title, MAP_HEADERS)

    src_ws = src_wb[args.src_sheet]
    src_positions = header_positions(src_ws)
    existing_rows = find_existing_map_rows(map_ws)
    imported_rows = 0
    updated_rows = 0

    for src_row in range(2, src_ws.max_row + 1):
        if not row_has_content(src_ws, src_row):
            continue

        values = {
            target: get_legacy_value(
                src_ws,
                src_row,
                src_positions,
                aliases,
                prefer="first" if target == "profile_url" else "last",
            )
            for target, aliases in LEGACY_MAP_COLUMNS.items()
        }
        if not values["sealib_id"] and not values["sealib_name"]:
            continue

        target_row = existing_rows.get(values["sealib_id"]) if values["sealib_id"] else None
        if target_row is None:
            target_row = map_ws.max_row + 1
            imported_rows += 1
        else:
            updated_rows += 1

        set_value(map_ws, target_row, "status", "imported")
        set_value(map_ws, target_row, "metric_source", args.source)
        set_value(map_ws, target_row, "metric_country", args.country)
        for target, value in values.items():
            if target == "sinta_journal_id" and not value:
                value = extract_journal_id(values["profile_url"])
            set_value(map_ws, target_row, target, value)

        current_confirmed = get_value(map_ws, target_row, "confirmed")
        if args.assume_confirmed and values["sinta_query_name"] and values["profile_url"]:
            set_value(map_ws, target_row, "confirmed", "1")
        elif current_confirmed == "":
            set_value(map_ws, target_row, "confirmed", "")

        if values["sealib_id"]:
            existing_rows[values["sealib_id"]] = target_row

    analysis_ws = src_wb[args.analysis_sheet]
    analysis_positions = header_positions(analysis_ws)
    existing_raw_keys = {
        (
            get_value(raw_ws, row, "journal_name"),
            get_value(raw_ws, row, "journal_id"),
            get_value(raw_ws, row, "profile_url"),
        )
        for row in range(2, raw_ws.max_row + 1)
    }
    imported_raw_rows = 0
    for src_row in range(2, analysis_ws.max_row + 1):
        if not row_has_content(analysis_ws, src_row):
            continue
        row_data = {
            header: get_legacy_value(analysis_ws, src_row, analysis_positions, [header])
            for header in RAW_HEADERS + DETAIL_HEADERS
        }
        if not any(row_data.values()):
            continue
        key = (row_data.get("journal_name", ""), row_data.get("journal_id", ""), row_data.get("profile_url", ""))
        if key in existing_raw_keys:
            continue
        target_row = raw_ws.max_row + 1
        for header in RAW_HEADERS + DETAIL_HEADERS:
            if header in row_data and row_data[header]:
                set_value(raw_ws, target_row, header, row_data[header])
        existing_raw_keys.add(key)
        imported_raw_rows += 1

    out_wb.save(args.xlsx)
    print(f"imported map rows: {imported_rows}, updated map rows: {updated_rows}")
    print(f"imported raw rows: {imported_raw_rows}")
    print(f"output: {args.xlsx}")
    return 0


def extract_json_payload(stdout: str) -> Any:
    text = stdout.strip()
    if not text:
        return []
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        start_candidates = [pos for pos in (text.find("["), text.find("{")) if pos >= 0]
        if not start_candidates:
            raise
        start = min(start_candidates)
        end = max(text.rfind("]"), text.rfind("}"))
        return json.loads(text[start : end + 1])


def normalize_candidates(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if isinstance(payload, dict):
        for key in ("results", "data", "items", "journals"):
            value = payload.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict)]
        return [payload]
    return []


def run_sinta_cli(args: argparse.Namespace, query: str) -> list[dict[str, Any]]:
    cmd_tokens = resolve_adapter_command(args)
    # Core contract: <python> <adapter-command> -q "<query>" -f json
    # SINTA profile flags (-m / --fetch-mode) are kept as the existing pass-through.
    cmd = [
        args.python,
        *cmd_tokens,
        "-q",
        query,
        "-m",
        args.mode,
        "-f",
        "json",
        "--fetch-mode",
        args.fetch_mode,
    ]
    completed = subprocess.run(
        cmd,
        check=False,
        text=True,
        capture_output=True,
        timeout=args.timeout,
    )
    if completed.returncode != 0:
        raise RuntimeError(completed.stderr.strip() or completed.stdout.strip())
    if completed.stderr.strip() and not completed.stdout.strip():
        raise RuntimeError(completed.stderr.strip())
    return normalize_candidates(extract_json_payload(completed.stdout))


def append_raw(ws, query: str, candidates: list[dict[str, Any]], fetched_at: str) -> None:
    headers = header_index(ws)
    for candidate in candidates:
        row = ws.max_row + 1
        set_value(ws, row, "query", query)
        set_value(ws, row, "fetched_at", fetched_at)
        for header in headers:
            if header in {"query", "fetched_at"}:
                continue
            set_value(ws, row, header, candidate.get(header, ""))


def should_fetch(ws, row: int, only: str | None, resume: bool) -> bool:
    if is_confirmed(get_value(ws, row, "confirmed")):
        return False
    status = get_value(ws, row, "status")
    if only and only != "all" and status != only:
        return False
    if not only and status not in {"", "needs_fetch", "needs_review", "manual_fixed", "fetch_error"}:
        return False
    if resume and get_value(ws, row, "fetched_at"):
        return False
    return True


def fetch_command(args: argparse.Namespace) -> int:
    wb = load_or_create_workbook(args.xlsx)
    raw_title, map_title = sheet_names(args.source, args.country)
    raw_headers = RAW_HEADERS + (DETAIL_HEADERS if args.fetch_mode == "detail" else [])
    raw_ws = ensure_sheet(wb, raw_title, raw_headers)
    map_ws = ensure_sheet(wb, map_title, MAP_HEADERS)

    resolve_adapter_command(args)  # fail fast if --adapter-command/--sinta-cli is missing

    processed = 0
    for row in range(2, map_ws.max_row + 1):
        if not should_fetch(map_ws, row, args.only, args.resume):
            continue
        query = get_value(map_ws, row, "sinta_query_name") or get_value(map_ws, row, "sealib_name")
        if not query:
            continue

        fetched_at = now_text()
        try:
            candidates = run_sinta_cli(args, query)
        except Exception as exc:
            set_value(map_ws, row, "status", "fetch_error")
            set_value(map_ws, row, "fetched_at", fetched_at)
            wb.save(args.xlsx)
            print(f"ERROR row {row} query={query}: {exc}", file=sys.stderr)
            processed += 1
            continue

        append_raw(raw_ws, query, candidates, fetched_at)
        if len(candidates) == 1:
            candidate = candidates[0]
            set_value(map_ws, row, "status", "auto_fetched")
            set_value(map_ws, row, "sinta_journal_name", candidate.get("journal_name", ""))
            set_value(map_ws, row, "sinta_journal_id", candidate.get("journal_id", ""))
            set_value(map_ws, row, "profile_url", candidate.get("profile_url", ""))
            set_value(map_ws, row, "sinta_level", candidate.get("sinta_level", ""))
            set_value(map_ws, row, "fetched_at", fetched_at)
        else:
            set_value(map_ws, row, "status", "needs_review")
            set_value(map_ws, row, "fetched_at", fetched_at)

        wb.save(args.xlsx)
        processed += 1
        if args.limit and processed >= args.limit:
            break

    print(f"fetch processed: {processed}")
    return 0


def refresh_command(args: argparse.Namespace) -> int:
    wb = load_or_create_workbook(args.xlsx)
    raw_title, map_title = sheet_names(args.source, args.country)
    raw_headers = RAW_HEADERS + (DETAIL_HEADERS if args.fetch_mode == "detail" else [])
    raw_ws = ensure_sheet(wb, raw_title, raw_headers)
    map_ws = ensure_sheet(wb, map_title, MAP_HEADERS)

    resolve_adapter_command(args)  # fail fast if --adapter-command/--sinta-cli is missing

    processed = 0
    matched = 0
    for row in range(2, map_ws.max_row + 1):
        if not is_confirmed(get_value(map_ws, row, "confirmed")):
            continue
        if args.resume and get_value(map_ws, row, "fetched_at").startswith(datetime.now().strftime("%Y-%m-%d")):
            continue

        journal_id = get_value(map_ws, row, "sinta_journal_id")
        query = get_value(map_ws, row, "sinta_query_name") or get_value(map_ws, row, "sealib_name")
        if not journal_id or not query:
            continue

        fetched_at = now_text()
        try:
            candidates = run_sinta_cli(args, query)
        except Exception as exc:
            print(f"ERROR row {row} query={query}: {exc}", file=sys.stderr)
            processed += 1
            continue

        append_raw(raw_ws, query, candidates, fetched_at)
        selected = next(
            (candidate for candidate in candidates if str(candidate.get("journal_id", "")).strip() == journal_id),
            None,
        )
        if selected:
            set_value(map_ws, row, "sinta_level", selected.get("sinta_level", ""))
            set_value(map_ws, row, "fetched_at", fetched_at)
            matched += 1
        else:
            set_value(map_ws, row, "status", "needs_review")
            set_value(map_ws, row, "fetched_at", fetched_at)
            print(f"WARNING row {row}: journal_id not found in refresh results: {journal_id}", file=sys.stderr)

        wb.save(args.xlsx)
        processed += 1
        if args.limit and processed >= args.limit:
            break

    print(f"refresh processed: {processed}, matched: {matched}")
    return 0


def export_command(args: argparse.Namespace) -> int:
    wb = load_or_create_workbook(args.xlsx)
    _, map_title = sheet_names(args.source, args.country)
    map_ws = ensure_sheet(wb, map_title, MAP_HEADERS)

    count = 0
    with args.out.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=EXPORT_HEADERS, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        for row in range(2, map_ws.max_row + 1):
            if not is_confirmed(get_value(map_ws, row, "confirmed")):
                continue
            grade, warning = normalized_grade(get_value(map_ws, row, "sinta_level"))
            if warning:
                print(f"{warning} (row {row})", file=sys.stderr)
            writer.writerow({
                "metric_source": get_value(map_ws, row, "metric_source") or args.source,
                "metric_country": get_value(map_ws, row, "metric_country") or args.country,
                "sealib_name": get_value(map_ws, row, "sealib_name"),
                "sealib_o_name": get_value(map_ws, row, "sealib_o_name"),
                "sealib_id": get_value(map_ws, row, "sealib_id"),
                "grade": grade,
                "url": get_value(map_ws, row, "profile_url"),
                "note": "",
            })
            count += 1

    print(f"exported rows: {count}")
    print(f"output: {args.out}")
    return 0


def report_command(args: argparse.Namespace) -> int:
    wb = load_or_create_workbook(args.xlsx)
    _, map_title = sheet_names(args.source, args.country)
    map_ws = ensure_sheet(wb, map_title, MAP_HEADERS)

    counts: Counter[str] = Counter()
    needs_review: list[tuple[int, str, str]] = []
    for row in range(2, map_ws.max_row + 1):
        status = get_value(map_ws, row, "status") or "(blank)"
        counts[status] += 1
        if status == "needs_review":
            needs_review.append((row, get_value(map_ws, row, "sealib_id"), get_value(map_ws, row, "sealib_name")))

    print("status counts:")
    for status, count in sorted(counts.items()):
        print(f"  {status}: {count}")
    print("\nneeds_review:")
    for row, sealib_id, name in needs_review:
        print(f"  row {row}\t{sealib_id}\t{name}")
    return 0


def add_common_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--source", default="SINTA")
    parser.add_argument("--country", default="ID")
    parser.add_argument("--xlsx", required=True, type=Path)


def add_sinta_args(parser: argparse.ArgumentParser) -> None:
    # Adapter invocation: --adapter-command is the script path (no default; explicit
    # is required for fetch/refresh). --sinta-cli is kept as a backward-compat alias.
    # The interpreter is taken from --python; do not pass an interpreter-prefixed command.
    parser.add_argument("--adapter-command")
    parser.add_argument("--sinta-cli", type=Path)
    parser.add_argument("--python", default=sys.executable)
    parser.add_argument("--mode", choices=["title", "all"], default="title")
    parser.add_argument("--fetch-mode", choices=["basic", "detail"], default="basic")
    parser.add_argument("--timeout", type=int, default=180)
    parser.add_argument("--resume", action="store_true")
    parser.add_argument("--limit", type=int, default=0)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build/update a Journal Metrics Excel correspondence table.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    init_parser = subparsers.add_parser("init")
    add_common_args(init_parser)
    init_parser.add_argument("--db", choices=["core", "ext"], default="ext")  # label only (補助)
    init_parser.add_argument("--db-path", required=True, type=Path)
    init_parser.add_argument("--sealib-country")
    init_parser.set_defaults(func=init_command)

    import_parser = subparsers.add_parser("import-existing-master")
    add_common_args(import_parser)
    import_parser.add_argument("--in", dest="input", required=True, type=Path)
    import_parser.add_argument("--src-sheet", default=DEFAULT_IMPORT_SRC_SHEET)
    import_parser.add_argument("--analysis-sheet", default=DEFAULT_IMPORT_ANALYSIS_SHEET)
    import_parser.add_argument("--assume-confirmed", action="store_true")
    import_parser.set_defaults(func=import_existing_master_command)

    fetch_parser = subparsers.add_parser("fetch")
    add_common_args(fetch_parser)
    add_sinta_args(fetch_parser)
    fetch_parser.add_argument("--only", choices=["all", "needs_fetch", "needs_review", "manual_fixed", "fetch_error"])
    fetch_parser.set_defaults(func=fetch_command)

    refresh_parser = subparsers.add_parser("refresh")
    add_common_args(refresh_parser)
    add_sinta_args(refresh_parser)
    refresh_parser.set_defaults(func=refresh_command)

    export_parser = subparsers.add_parser("export")
    add_common_args(export_parser)
    export_parser.add_argument("--out", required=True, type=Path)
    export_parser.set_defaults(func=export_command)

    report_parser = subparsers.add_parser("report")
    add_common_args(report_parser)
    report_parser.set_defaults(func=report_command)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
