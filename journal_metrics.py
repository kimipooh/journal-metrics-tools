#!/usr/bin/env python3
"""Journal Metrics Workflow CLI."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any, Iterable, Sequence

from adapters.mock import fetch_journal as fetch_mock_journal
from adapters.sealib import fetch_journal as fetch_sealib_journal
from journal_mapper import map_envelope_to_journal_rows
from journal_mapper import row_to_journal_values
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


MAIN_HEADERS = [
    "id",
    "name",
    "o_name",
    "issn",
    "eissn",
    "journal_name",
    "note",
    "status",
]

JOURNAL_HEADERS = [
    "main_row_id",
    "journal_type",
    "external_journal_id",
    "journal_name",
    "affiliation",
    "grade",
    "profile_url",
    "fetch_status",
    "fetched_at",
    "raw_json",
]

MAIN_STATUS_BY_ENVELOPE_STATUS = {
    "fetched": "fetched",
    "multiple_candidates": "multiple_candidates",
    "not_found": "not_found",
    "adapter_error": "adapter_error",
}

CONVERT_HEADERS = [
    "main_row_id",
    "metric_source",
    "metric_country",
    "sealib_name",
    "sealib_o_name",
    "sealib_id",
    "grade",
    "url",
    "note",
    "convert_status",
]

PROGRAM2_TSV_HEADERS = [
    "metric_source",
    "metric_country",
    "sealib_name",
    "sealib_o_name",
    "sealib_id",
    "grade",
    "url",
    "note",
]

README_ROWS = [
    ("Section", "Description"),
    ("Purpose", "Journal Metrics Workflow Phase 1 template workbook."),
    ("Usage", "python journal_metrics.py template --output journal_metrics.xlsx"),
    ("main", "Human-edited input sheet. The status column is created but not processed automatically."),
    ("journal", "Placeholder for fetched journal candidates in later phases."),
    ("convert", "Placeholder for confirmed rows prepared for later export or database import."),
    ("Phase 1 limits", "No external CLI, database, adapter, fetch-journal, convert, or enrich-db is called."),
]

DEFAULT_WIDTH = 14
MIN_WIDTH_PADDING = 2
MAX_WIDTH = 48


def apply_header_style(ws: Worksheet) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def set_reasonable_widths(ws: Worksheet, rows: Iterable[Sequence[object]]) -> None:
    widths: dict[int, int] = {}
    for row in rows:
        for index, value in enumerate(row, start=1):
            width = len(str(value)) + MIN_WIDTH_PADDING if value is not None else DEFAULT_WIDTH
            widths[index] = max(widths.get(index, DEFAULT_WIDTH), width)

    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = min(width, MAX_WIDTH)


def write_rows(ws: Worksheet, rows: Sequence[Sequence[object]]) -> None:
    for row in rows:
        ws.append(row)
    apply_header_style(ws)
    set_reasonable_widths(ws, rows)


def write_header_sheet(ws: Worksheet, headers: Sequence[str]) -> None:
    write_rows(ws, [headers])


def append_journal_rows(
    ws: Worksheet,
    rows: list[dict],
    headers: list[str],
) -> int:
    for row in rows:
        ws.append(row_to_journal_values(row, headers))
    return len(rows)


def append_convert_rows(
    ws: Worksheet,
    rows: list[dict],
    headers: list[str],
) -> int:
    for row in rows:
        ws.append([row.get(header) for header in headers])
    return len(rows)


def reset_convert_sheet(ws: Worksheet) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    if ws.max_column > len(CONVERT_HEADERS):
        ws.delete_cols(len(CONVERT_HEADERS) + 1, ws.max_column - len(CONVERT_HEADERS))
    for column, header in enumerate(CONVERT_HEADERS, start=1):
        ws.cell(row=1, column=column, value=header)
    apply_header_style(ws)
    set_reasonable_widths(ws, [CONVERT_HEADERS])


def header_index(ws: Worksheet) -> dict[str, int]:
    return {
        str(cell.value): index
        for index, cell in enumerate(ws[1])
        if cell.value is not None
    }


def cell_value(row: Sequence[object], index: dict[str, int], header: str) -> object:
    column_index = index.get(header)
    if column_index is None:
        return None
    if column_index >= len(row):
        return None
    return row[column_index].value


def text_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def row_to_dict(headers: list[str], values: Sequence[object]) -> dict[str, object]:
    return {
        header: values[index] if index < len(values) else None
        for index, header in enumerate(headers)
    }


def tsv_value(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def note_looks_like_json(value: str) -> bool:
    text = value.strip()
    if not text:
        return False
    return (
        (text.startswith("{") and text.endswith("}"))
        or (text.startswith("[") and text.endswith("]"))
    )


def note_has_key_value_format(value: str) -> bool:
    text = value.strip()
    if not text:
        return True
    parts = [part.strip() for part in text.split(";")]
    for part in parts:
        if not part or "=" not in part:
            return False
        key, _value = part.split("=", 1)
        if not key.strip():
            return False
    return True


def normalized_main_row_id(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = text_value(value)
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        return None


def nullable_text(value: object) -> str | None:
    text = text_value(value)
    return text or None


def raw_json_object(value: object) -> dict[str, Any]:
    if not isinstance(value, str) or not value.strip():
        return {}
    try:
        parsed = json.loads(value)
    except json.JSONDecodeError:
        return {}
    if not isinstance(parsed, dict):
        return {}
    return parsed


def extract_metric_country(journal_row: dict) -> str | None:
    raw_candidate = raw_json_object(journal_row.get("raw_json"))
    return nullable_text(raw_candidate.get("country"))


def build_note(
    *,
    external_journal_id: object,
    publisher: object,
    eissn: object,
) -> str | None:
    parts = []
    for key, value in [
        ("external_id", external_journal_id),
        ("affiliation", publisher),
        ("eissn", eissn),
    ]:
        text = nullable_text(value)
        if text is not None:
            parts.append(f"{key}={text}")
    if not parts:
        return None
    return "; ".join(parts)


def generate_convert_rows(
    main_row: dict,
    journal_row: dict,
) -> dict | None:
    """Build one convert row from an accepted journal row."""
    if text_value(journal_row.get("fetch_status")).lower() != "ok":
        return None

    raw_candidate = raw_json_object(journal_row.get("raw_json"))

    return {
        "main_row_id": journal_row.get("main_row_id"),
        "metric_source": journal_row.get("journal_type"),
        "metric_country": extract_metric_country(journal_row),
        # TODO: Fill these from enrich-db/main once SEALIB enrichment is implemented.
        "sealib_name": None,
        "sealib_o_name": None,
        "sealib_id": None,
        "grade": journal_row.get("grade"),
        "url": journal_row.get("profile_url"),
        "note": build_note(
            external_journal_id=raw_candidate.get("external_journal_id"),
            publisher=raw_candidate.get("publisher"),
            eissn=raw_candidate.get("eissn"),
        ),
        "convert_status": "ready",
    }


def build_template_workbook() -> Workbook:
    wb = Workbook()

    readme = wb.active
    readme.title = "README"
    write_rows(readme, README_ROWS)

    main = wb.create_sheet("main")
    write_header_sheet(main, MAIN_HEADERS)

    journal = wb.create_sheet("journal")
    write_header_sheet(journal, JOURNAL_HEADERS)

    convert = wb.create_sheet("convert")
    write_header_sheet(convert, CONVERT_HEADERS)

    return wb


def template_command(args: argparse.Namespace) -> None:
    output_path = Path(args.output)
    wb = build_template_workbook()
    wb.save(output_path)
    print(f"Created template workbook: {output_path}")


def fetch_journal_command(args: argparse.Namespace) -> None:
    if args.adapter == "sealib" and not args.db_path:
        raise SystemExit("ERROR: --db-path is required when --adapter sealib")

    input_path = Path(args.input)
    wb = load_workbook(input_path)
    if "main" not in wb.sheetnames:
        raise ValueError("Workbook does not contain a main sheet")
    if "journal" not in wb.sheetnames:
        raise ValueError("Workbook does not contain a journal sheet")

    main_ws = wb["main"]
    journal_ws = wb["journal"]
    main_headers = header_index(main_ws)

    for required_header in ["status", "journal_name", "name"]:
        if required_header not in main_headers:
            raise ValueError(f"main sheet is missing required header: {required_header}")

    processed_rows = 0
    appended_rows = 0

    for excel_row_number, row in enumerate(
        main_ws.iter_rows(min_row=2),
        start=2,
    ):
        status = text_value(cell_value(row, main_headers, "status")).lower()
        if status not in {"", "pending"}:
            continue

        query = text_value(cell_value(row, main_headers, "journal_name"))
        if not query:
            query = text_value(cell_value(row, main_headers, "name"))
        if not query:
            continue

        if args.adapter == "mock":
            envelope = fetch_mock_journal(query)
        elif args.adapter == "sealib":
            envelope = fetch_sealib_journal(
                query,
                db_path=args.db_path,
                country=args.country,
            )
        else:
            raise ValueError(f"Unsupported adapter: {args.adapter}")

        journal_rows = map_envelope_to_journal_rows(
            envelope,
            main_row_id=excel_row_number,
        )
        appended_rows += append_journal_rows(
            journal_ws,
            journal_rows,
            JOURNAL_HEADERS,
        )
        main_ws.cell(
            row=excel_row_number,
            column=main_headers["status"] + 1,
            value=MAIN_STATUS_BY_ENVELOPE_STATUS[envelope["status"]],
        )
        processed_rows += 1

    wb.save(input_path)
    print(f"Processed main rows: {processed_rows}")
    print(f"Appended journal rows: {appended_rows}")
    print(f"Adapter: {args.adapter}")


def convert_command(args: argparse.Namespace) -> None:
    input_path = Path(args.input)
    wb = load_workbook(input_path)
    for sheet_name in ["main", "journal", "convert"]:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Workbook does not contain a {sheet_name} sheet")

    main_ws = wb["main"]
    journal_ws = wb["journal"]
    convert_ws = wb["convert"]

    main_headers = [str(cell.value) for cell in main_ws[1] if cell.value is not None]
    journal_headers = [str(cell.value) for cell in journal_ws[1] if cell.value is not None]

    main_rows_by_id: dict[int, dict] = {}
    for excel_row_number, values in enumerate(
        main_ws.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        main_row = row_to_dict(main_headers, values)
        main_row["main_row_id"] = excel_row_number
        main_rows_by_id[excel_row_number] = main_row

    reset_convert_sheet(convert_ws)

    processed_journal_rows = 0
    appended_convert_rows = 0
    for values in journal_ws.iter_rows(min_row=2, values_only=True):
        processed_journal_rows += 1
        journal_row = row_to_dict(journal_headers, values)
        main_row_id = normalized_main_row_id(journal_row.get("main_row_id"))
        if main_row_id is None:
            continue
        main_row = main_rows_by_id.get(main_row_id)
        if main_row is None:
            continue
        convert_row = generate_convert_rows(main_row, journal_row)
        if convert_row is None:
            continue
        appended_convert_rows += append_convert_rows(
            convert_ws,
            [convert_row],
            CONVERT_HEADERS,
        )

    wb.save(input_path)
    print(f"processed_journal_rows = {processed_journal_rows}")
    print(f"appended_convert_rows = {appended_convert_rows}")


def export_tsv_command(args: argparse.Namespace) -> None:
    input_path = Path(args.input)
    output_path = Path(args.output)
    wb = load_workbook(input_path, read_only=True)
    if "convert" not in wb.sheetnames:
        raise ValueError("Workbook does not contain a convert sheet")

    convert_ws = wb["convert"]
    convert_headers = [str(cell.value) for cell in convert_ws[1] if cell.value is not None]
    missing_headers = [header for header in PROGRAM2_TSV_HEADERS + ["convert_status"] if header not in convert_headers]
    if missing_headers:
        raise ValueError(f"convert sheet is missing required headers: {', '.join(missing_headers)}")

    exported_rows = 0
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
        writer.writerow(PROGRAM2_TSV_HEADERS)
        for values in convert_ws.iter_rows(min_row=2, values_only=True):
            convert_row = row_to_dict(convert_headers, values)
            if text_value(convert_row.get("convert_status")).lower() != "ready":
                continue
            writer.writerow([
                tsv_value(convert_row.get(header))
                for header in PROGRAM2_TSV_HEADERS
            ])
            exported_rows += 1

    print(f"exported_tsv_rows = {exported_rows}")
    print(f"output = {output_path}")


def validate_tsv_command(args: argparse.Namespace) -> None:
    input_path = Path(args.input)
    errors: list[str] = []
    warnings: list[str] = []
    rows = 0

    with input_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle, delimiter="\t")
        try:
            headers = next(reader)
        except StopIteration:
            errors.append("header row is missing")
            headers = []

        if headers and headers != PROGRAM2_TSV_HEADERS:
            errors.append(
                "header must exactly match: "
                + "\t".join(PROGRAM2_TSV_HEADERS)
            )

        for line_number, row in enumerate(reader, start=2):
            rows += 1
            if len(row) != len(PROGRAM2_TSV_HEADERS):
                errors.append(
                    f"row {line_number}: expected {len(PROGRAM2_TSV_HEADERS)} columns, got {len(row)}"
                )
                continue

            row_data = dict(zip(PROGRAM2_TSV_HEADERS, row, strict=True))
            for required_header in ["metric_source", "metric_country", "grade"]:
                if not text_value(row_data.get(required_header)):
                    errors.append(
                        f"row {line_number}: {required_header} is required"
                    )

            if not any(
                text_value(row_data.get(header))
                for header in ["sealib_name", "sealib_o_name", "sealib_id"]
            ):
                errors.append(
                    f"row {line_number}: one of sealib_name, sealib_o_name, sealib_id is required"
                )

            note = text_value(row_data.get("note"))
            if note:
                if note_looks_like_json(note):
                    warnings.append(
                        f"row {line_number}: note looks like raw JSON"
                    )
                elif not note_has_key_value_format(note):
                    warnings.append(
                        f"row {line_number}: note should use key=value; key=value format"
                    )

    for error in errors:
        print(f"ERROR: {error}")
    for warning in warnings:
        print(f"WARNING: {warning}")
    print(f"rows = {rows}")
    print(f"errors = {len(errors)}")
    print(f"warnings = {len(warnings)}")

    if errors:
        raise SystemExit(1)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Journal Metrics Workflow tools.",
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    template = subparsers.add_parser(
        "template",
        help="Create a Phase 1 Journal Metrics template workbook.",
    )
    template.add_argument(
        "--output",
        required=True,
        help="Output .xlsx path. Existing files are overwritten.",
    )
    template.set_defaults(func=template_command)

    fetch_journal = subparsers.add_parser(
        "fetch-journal",
        help="Fetch journal candidates into the journal sheet.",
    )
    fetch_journal.add_argument(
        "--input",
        required=True,
        help="Input .xlsx path to update in place.",
    )
    fetch_journal.add_argument(
        "--adapter",
        required=True,
        choices=["mock", "sealib"],
        help="Adapter to use.",
    )
    fetch_journal.add_argument(
        "--db-path",
        help="SEALIB SQLite DB path. Required when --adapter sealib.",
    )
    fetch_journal.add_argument(
        "--country",
        help="Optional SEALIB country filter used only with --adapter sealib.",
    )
    fetch_journal.set_defaults(func=fetch_journal_command)

    convert = subparsers.add_parser(
        "convert",
        help="Generate convert sheet rows from accepted journal rows.",
    )
    convert.add_argument(
        "--input",
        required=True,
        help="Input .xlsx path to update in place.",
    )
    convert.set_defaults(func=convert_command)

    export_tsv = subparsers.add_parser(
        "export-tsv",
        help="Export ready convert rows as a Program2 TSV file.",
    )
    export_tsv.add_argument(
        "--input",
        required=True,
        help="Input .xlsx path to read.",
    )
    export_tsv.add_argument(
        "--output",
        required=True,
        help="Output .tsv path. Existing files are overwritten.",
    )
    export_tsv.set_defaults(func=export_tsv_command)

    validate_tsv = subparsers.add_parser(
        "validate-tsv",
        help="Validate a Program2 TSV file without writing to a database.",
    )
    validate_tsv.add_argument(
        "--input",
        required=True,
        help="Input .tsv path to validate.",
    )
    validate_tsv.set_defaults(func=validate_tsv_command)

    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    args.func(args)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
